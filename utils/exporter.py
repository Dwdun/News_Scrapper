import csv
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# FIELDS: bagian data artikel yang mau dimasukkan ke file
FIELDS = ["title", "date", "content", "url"]

# HEADERS: judul kolom yang akan tampil di bagian atas file export (termasuk No)
HEADERS = ["No", "Judul", "Tanggal", "Isi", "URL"]


def _clean_text(text):
    """Membersihkan teks dari newline dan spasi berlebih agar rapi di tabel."""
    if not text:
        return ""
    # Ganti semua jenis newline dan tab menjadi spasi tunggal
    cleaned = re.sub(r'[\r\n\t]+', ' ', str(text))
    # Hilangkan spasi berlebihan (dobel atau lebih)
    cleaned = re.sub(r' {2,}', ' ', cleaned)
    return cleaned.strip()


def _format_date(date_str):
    """Konversi tanggal ISO (2026-03-08 00:00:00) ke format Indonesia (DD/MM/YYYY HH.mm).
    Jika bukan format ISO, kembalikan apa adanya.
    """
    if not date_str:
        return ""
    # Coba parse format ISO datetime: YYYY-MM-DD HH:MM:SS atau YYYY-MM-DD
    iso_pattern = re.match(r'(\d{4})-(\d{2})-(\d{2})(?:[T ]?(\d{2}):(\d{2}))?', str(date_str))
    if iso_pattern:
        year = iso_pattern.group(1)
        month = iso_pattern.group(2)
        day = iso_pattern.group(3)
        hour = iso_pattern.group(4) or '00'
        minute = iso_pattern.group(5) or '00'
        return f"{day}/{month}/{year} {hour}.{minute}"
    # Bukan format ISO, kembalikan apa adanya (misal: '3 tahun lalu', '05-Mar-26')
    return str(date_str).strip()


def export_to_csv(data, filepath, content_limit=0):
    """
    Mengekspor daftar artikel ke dalam file format CSV yang rapi.
    Menggunakan delimiter titik koma (;) agar langsung terbuka rapi di Excel.
    content_limit: batas karakter isi (0 = tampilkan semua).
    """
    try:
        # Membuka file CSV untuk ditulis
        with open(filepath, "w", newline="", encoding="utf-8-sig") as file:
            # Gunakan delimiter ; agar kolom terpisah otomatis di Excel
            writer = csv.writer(file, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            
            # Menulis baris pertama dengan header yang rapi
            writer.writerow(HEADERS)
            
            # Menambahkan setiap artikel ke dalam file CSV
            for i, row in enumerate(data, 1):
                # Ambil dan bersihkan isi content
                content = _clean_text(row.get("content", ""))
                # Terapkan limit karakter jika di-set
                if content_limit > 0 and len(content) > content_limit:
                    content = content[:content_limit] + "..."

                writer.writerow([
                    i,                                          # No
                    _clean_text(row.get("title", "")),          # Judul
                    _format_date(row.get("date", "")),          # Tanggal
                    content,                                    # Isi (sesuai limit)
                    row.get("url", "")                          # URL
                ])
                
        # File berhasil diexport
        return True
    
    except Exception:
        # Jika ada file yang gagal diekspor (error)
        return False

def export_to_excel(data, filepath, content_limit=0):
    """
    Mengekspor daftar artikel ke dalam file format Excel yang rapi.
    Lebar kolom tetap: No=6, Judul=90, Tanggal=30, Isi=180, URL=180.
    Header berwarna biru + bold. Isi menggunakan wrap text.
    content_limit: batas karakter isi (0 = tampilkan semua).
    """
    try:
        # Membuat file Excel baru
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hasil Scraping"
        
        # ============================
        # HEADER — baris pertama
        # ============================
        ws.append(HEADERS)
        
        # Style header: bold, putih, latar biru
        header_font = Font(name='Segoe UI', bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # ============================
        # LEBAR KOLOM (tetap)
        # ============================
        ws.column_dimensions['A'].width = 6     # No
        ws.column_dimensions['B'].width = 90    # Judul
        ws.column_dimensions['C'].width = 30    # Tanggal
        ws.column_dimensions['D'].width = 180   # Isi
        ws.column_dimensions['E'].width = 180   # URL
        
        # ============================
        # STYLE DATA
        # ============================
        font_data = Font(name='Segoe UI', size=10)
        center_top = Alignment(horizontal="center", vertical="top")
        left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
        left_nowrap = Alignment(horizontal="left", vertical="top", wrap_text=False)
        even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
            
        # ============================
        # DATA — baris ke-2 dst
        # ============================
        for i, row in enumerate(data, 1):
            # Ambil dan bersihkan isi content
            content = _clean_text(row.get("content", ""))
            # Terapkan limit karakter jika di-set
            if content_limit > 0 and len(content) > content_limit:
                content = content[:content_limit] + "..."

            ws.append([
                i,                                          # No
                _clean_text(row.get("title", "")),          # Judul
                _format_date(row.get("date", "")),          # Tanggal
                content,                                    # Isi (sesuai limit)
                row.get("url", "")                          # URL
            ])
            
            r = i + 1  # baris di Excel (1 = header)
            
            # No — center
            ws.cell(row=r, column=1).alignment = center_top
            ws.cell(row=r, column=1).font = font_data
            # Judul — left wrap
            ws.cell(row=r, column=2).alignment = left_wrap
            ws.cell(row=r, column=2).font = font_data
            # Tanggal — center
            ws.cell(row=r, column=3).alignment = center_top
            ws.cell(row=r, column=3).font = font_data
            # Isi — left wrap (teks panjang akan melebar ke bawah)
            ws.cell(row=r, column=4).alignment = left_wrap
            ws.cell(row=r, column=4).font = font_data
            # URL — left no wrap
            ws.cell(row=r, column=5).alignment = left_nowrap
            ws.cell(row=r, column=5).font = font_data
            
            # Warna selang-seling (baris genap)
            if i % 2 == 0:
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = even_fill

        # Freeze header
        ws.freeze_panes = 'A2'

        # Simpan file
        wb.save(filepath)
        return True
            
    except Exception as e:
        return str(e)

